"""
Excel Renderer — converts structured data to .xlsx via openpyxl.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def run(input_data) -> dict:
    """
    Render structured data to an .xlsx file.

    Args:
        input_data: dict with "sheets" list and "output_path"
        Each sheet: {"name": "Sheet1", "headers": [...], "rows": [[...], ...]}
    """
    if not isinstance(input_data, dict):
        return {"status": "error", "result": "Invalid input: expected dict"}

    sheets = input_data.get("sheets", [])
    output_path = input_data.get("output_path", "outputs/data.xlsx")

    if not sheets:
        return {"status": "error", "result": "No sheets provided"}

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    for sheet_data in sheets:
        ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, len(rows) + 2)
            )
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return {
        "status": "ok",
        "result": f"Excel file saved: {output_path}",
        "files": [output_path],
    }
